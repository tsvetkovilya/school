import os
from urllib.parse import quote

import magic
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.db import IntegrityError
from django.http import HttpResponse, HttpResponseNotFound, FileResponse
from openpyxl import Workbook
from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.models import User
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from school.models import Teacher, CustomUser
from schoolsite import settings
from .forms import RegisterForm, LoginForm
from .models import Course, CourseRegistration, Lesson, Counter, Attendance_of_classes, Attendance


def register(request):
    if request.method == 'GET':
        reg = RegisterForm()
        context = {'reg': reg}
        return render(request, 'register.html', context)
    if request.method == 'POST':
        reg = RegisterForm(request.POST)
    if reg.is_valid():
        user = reg.save(commit=False)  # Сохраняем объект модели User
        user.wishes = reg.cleaned_data['wishes']  # Добавляем дополнительную информацию
        user.phone_number = reg.cleaned_data['phone_number']
        user.save()
        new_user = authenticate(username=reg.cleaned_data['username'],
                                password=reg.cleaned_data['password1'],
                                )
        login(request, new_user)
        return redirect('Home')
    else:
        print('Form is not valid')
        messages.error(request, 'Error Processing Your Request')
        context = {'reg': reg}
        return render(request, 'register.html', context)
    return render(request, 'register.html', {'reg': reg})

def login_view(request):
    if request.method == 'POST':
        form = LoginForm(request, request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('Home')
            else:
                form.add_error(None, "Invalid username or password.")
    else:
        form = LoginForm()

    return render(request, 'login.html', {'form': form})

def logout_view(request):
    logout(request)
    return redirect('Home')

def course_registration(request):

    if request.method == 'POST':

        if request.user.is_authenticated:
            student_id = request.user.id
        else:
            return redirect('Login')

        course_id = request.POST.get('course_id')


        print(student_id)
        print(course_id)

        try:
            registration = CourseRegistration(course_id=course_id, student_id=student_id)
            registration.save()
            print('Запись прошла')
            messages.success(request, 'You have been successfully registered for the course.')
        except ValidationError as e:
            messages.warning(request, 'This student is already registered for this course.')

        return redirect('courses')
    else:
        courses = Course.objects.all()
        counter = Counter()
        courses_reg = CourseRegistration.objects.all()
        return render(request, 'courses.html', {'courses': courses, 'courses_reg': courses_reg, 'counter': counter})

@login_required
def profile(request):
    counter = Counter()
    courses = CourseRegistration.objects.filter(student_id=request.user.id)
    courses_names = Course.objects.all()
    teachers = Teacher.objects.all()
    lessons = Lesson.objects.order_by('start_time')

    return render(request, 'account.html', {'courses': courses, 'courses_names': courses_names, 'teachers': teachers, 'lessons': lessons, 'counter': counter})

def schedule(request):
    lessons = Lesson.objects.order_by('start_time')
    return render(request, 'account.html', {'lessons': lessons})


def download_file(request, file_name):
    file_path = os.path.join(settings.MEDIA_ROOT, file_name)

    if os.path.exists(file_path):
        return FileResponse(open(file_path, 'rb'), content_type='application/octet-stream')
    else:
        return HttpResponseNotFound('Файл не найден.')

def students_courses_report(request):
    students = CustomUser.objects.all()
    courses_registration = CourseRegistration.objects.all()
    courses = Course.objects.all()
    student_course_data = []

    for course in courses:
        for student in students:
            for course_reg in courses_registration:
                if student.is_staff == 0:
                    if student.id == course_reg.student_id and course_reg.course_id == course.id:
                        dj = course_reg.created_at
                        dj = dj.replace(tzinfo=None)
                        student_course_data.append({
                            'student_id': student.id,
                            'student': student.last_name + ' ' + student.first_name,
                            'course_name': course.name,
                            'created_at': dj
                        })

    wb = Workbook()
    ws = wb.active
    ws.append(['ID студента', 'Cтудент', 'Название курса', 'Дата регистрации на курс'])
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    for data in student_course_data:
        ws.append([data['student_id'], data['student'], data['course_name'], data['created_at']])

    ws.auto_filter.ref = 'A1:D1'

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="list_of_students_and_their_course_entries.xlsx"'
    wb.save(response)

    return response

def students_report(request):
    students = CustomUser.objects.all()
    courses_registration = CourseRegistration.objects.all()
    courses = Course.objects.all()
    student_course_data = []

    for student in students:
        if student.is_staff == 0:
            dj = student.date_joined
            dj = dj.replace(tzinfo=None)
            student_course_data.append({
                'student_id': student.id,
                'student_username': student.username,
                'student': student.last_name + ' ' + student.first_name,
                'student_email': student.email,
                'student_date_joined': dj,
                'student_wishes': student.wishes,
                'student_phone': student.phone_number,
                'student_grade': student.grade
            })

    wb = Workbook()
    ws = wb.active
    ws.append(['ID студента', 'Никнейм', 'Фамилия и имя', 'Почта студента', 'Дата регистрации', 'Пожелания к обучению', 'Номер студента', 'Класс в котором учится'])
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 23

    for data in student_course_data:
        ws.append([data['student_id'], data['student_username'], data['student'], data['student_email'], data['student_date_joined'], data['student_wishes'], data['student_phone'], data['student_grade']])

    ws.auto_filter.ref = 'A1:H1'

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="list_of_students.xlsx"'
    wb.save(response)

    return response

def attendance_and_classes_report(request):
    attendances = Attendance_of_classes.objects.all()
    attendance_and_classes = []

    for attendance in attendances:
        enrolled_students = ", ".join([student.student.last_name + ' ' + student.student.first_name for student in attendance.lesson.students.all()])
        visiting_students = ", ".join([student.student.last_name + ' ' + student.student.first_name for student in attendance.students.all()])

        # for enrolled in attendance.lesson.students.all():
        #     for visiting in attendance.students.all():
        #         if enrolled.id == visiting.id:
        #             # absent_students = ''.join(enrolled.last_name)

        date_of_the_lesson = f"{attendance.lesson.day_of_week.strftime('%Y-%m-%d')} {attendance.lesson.start_time}"

        attendance_and_classes.append({
            'course_name': attendance.lesson.course.name,
            'lesson_topic': attendance.lesson.lesson_topic,
            'teacher': attendance.lesson.teacher.teacher.last_name + ' ' + attendance.lesson.teacher.teacher.first_name + ' ' + attendance.lesson.teacher.patronymic,
            'date_of_the_lesson': date_of_the_lesson,
            'enrolled_students': enrolled_students,
            'visiting_students': visiting_students,
            # 'absent_students': absent_students
        })

    wb = Workbook()
    ws = wb.active
    ws.append(['Название курса', 'Тема занятия', 'Дата занятия', 'Преподаватель', 'Записанные студенты', 'Был на занятии(+), не был на занятии(-)'])
    column_widths = [18, 25, 25, 32, 25, 40]

    for i, width in enumerate(column_widths, start=1):
        column_letter = get_column_letter(i)
        ws.column_dimensions[column_letter].width = width

    for data in attendance_and_classes:
        enrolled_students = data['enrolled_students'].split(', ')
        visiting_students = data['visiting_students'].split(', ')

        max_len = max(len(enrolled_students), len(visiting_students))

        for i in range(max_len):
            enrolled = enrolled_students[i] if i < len(enrolled_students) else ''
            visiting = ''

            if enrolled and enrolled in enrolled_students:
                visiting += "+" if enrolled in visiting_students else "-"

            if i == 0:
                ws.append([data['course_name'], data['lesson_topic'], data['date_of_the_lesson'], data['teacher'], enrolled, visiting])
            else:
                ws.append(['', '', '', '', enrolled, visiting])

    ws.auto_filter.ref = 'A1:F1'

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="attendance_and_classes.xlsx"'
    wb.save(response)

    return response

def student_performance(request):
    attendances = Attendance_of_classes.objects.all()
    attendance_and_classes = []
    grades = Attendance.objects.all()

    for attendance in attendances:
        enrolled_students = ", ".join(
            [student.student.last_name + ' ' + student.student.first_name for student in attendance.lesson.students.all()])
        visiting_students = ", ".join(
            [student.student.last_name + ' ' + student.student.first_name for student in attendance.students.all()])

        grade_students = {}  # Создаем словарь для хранения оценок

        for grade in grades:
            if grade.attendance_id == attendance.id:
                student_name = grade.student.last_name + ' ' + grade.student.first_name
                grade_students[student_name] = grade.grade  # Добавляем оценку в словарь

        date_of_the_lesson = f"{attendance.lesson.day_of_week.strftime('%Y-%m-%d')} {attendance.lesson.start_time}"

        attendance_and_classes.append({
            'course_name': attendance.lesson.course.name,
            'lesson_topic': attendance.lesson.lesson_topic,
            'teacher': attendance.lesson.teacher.teacher.last_name + ' ' + attendance.lesson.teacher.teacher.first_name + ' ' + attendance.lesson.teacher.patronymic,
            'date_of_the_lesson': date_of_the_lesson,
            'enrolled_students': enrolled_students,
            'visiting_students': visiting_students,
            'grade_students': grade_students  # Добавляем словарь оценок в словарь `attendance_and_classes`
        })

    wb = Workbook()
    ws = wb.active
    ws.append(['Название курса', 'Тема занятия', 'Дата занятия', 'Преподаватель', 'Записанные студенты',
               'Был на занятии(+), не был на занятии(-)', 'Оценка'])
    column_widths = [18, 25, 25, 32, 25, 40, 10]

    for i, width in enumerate(column_widths, start=1):
        column_letter = get_column_letter(i)
        ws.column_dimensions[column_letter].width = width

    for data in attendance_and_classes:
        enrolled_students = data['enrolled_students'].split(', ')
        visiting_students = data['visiting_students'].split(', ')

        max_len = max(len(enrolled_students), len(visiting_students))

        for i in range(max_len):
            enrolled = enrolled_students[i] if i < len(enrolled_students) else ''
            visiting = ''

            if enrolled and enrolled in enrolled_students:
                visiting += "+" if enrolled in visiting_students else "-"

            gradestudents = data['grade_students'].get(enrolled, '')  # Получаем оценку из словаря grade_students

            if i == 0:
                ws.append(
                    [data['course_name'], data['lesson_topic'], data['date_of_the_lesson'], data['teacher'], enrolled,
                     visiting, gradestudents])
            else:
                ws.append(['', '', '', '', enrolled, visiting, gradestudents])

    ws.auto_filter.ref = 'A1:G1'

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="student_performance.xlsx"'
    wb.save(response)

    return response